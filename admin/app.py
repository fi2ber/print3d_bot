import os
import sys
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import openpyxl
from io import BytesIO

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))

from config import Config
from database.models import Database

app = Flask(__name__)
app.config['SECRET_KEY'] = Config.SECRET_KEY
app.config['SQLALCHEMY_DATABASE_URI'] = Config.DATABASE_URL
app.config['UPLOAD_FOLDER'] = Config.UPLOAD_FOLDER

db_flask = SQLAlchemy(app)

# Initialize database
db = Database(Config.DATABASE_URL)

@app.route('/')
def index():
    """Redirect to admin panel"""
    return redirect(url_for('admin'))

@app.route('/admin')
def admin():
    """Admin panel main page"""
    try:
        # Get filter parameters
        status_filter = request.args.get('status', 'all')
        search = request.args.get('search', '')
        
        # Get all orders
        if status_filter == 'all':
            orders = db.get_all_orders()
        else:
            orders = db.get_orders_by_status(status_filter)
        
        # Apply search filter
        if search:
            orders = [order for order in orders if 
                     search.lower() in order.name.lower() or 
                     search.lower() in str(order.id) or
                     search.lower() in order.phone.lower()]
        
        # Convert to dict for template
        orders_data = []
        for order in orders:
            orders_data.append({
                'id': order.id,
                'name': order.name,
                'phone': order.phone,
                'username': order.username,
                'service_type': Config.SERVICE_TYPES.get(order.service_type, order.service_type),
                'status': order.status,
                'status_display': Config.ORDER_STATUSES.get(order.status, order.status),
                'material': order.material,
                'color': order.color,
                'quantity': order.quantity,
                'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
                'description': order.description,
                'file_path': order.file_path,
                'language': order.language,
                'manager_contact': order.manager_contact,
                'location': order.location
            })
        
        return render_template('admin.html', 
                             orders=orders_data,
                             statuses=Config.ORDER_STATUSES,
                             service_types=Config.SERVICE_TYPES,
                             status_filter=status_filter,
                             search=search)
    
    except Exception as e:
        flash(f'Ошибка при загрузке заказов: {str(e)}', 'error')
        return render_template('admin.html', orders=[], statuses={}, service_types={})

@app.route('/admin/order/<int:order_id>')
def order_detail(order_id):
    """Order detail page"""
    order = db.get_order(order_id)
    if not order:
        flash('Заказ не найден', 'error')
        return redirect(url_for('admin'))
    
    order_data = {
        'id': order.id,
        'name': order.name,
        'phone': order.phone,
        'username': order.username,
        'service_type': Config.SERVICE_TYPES.get(order.service_type, order.service_type),
        'status': order.status,
        'status_display': Config.ORDER_STATUSES.get(order.status, order.status),
        'material': order.material,
        'color': order.color,
        'strength_accuracy': order.strength_accuracy,
        'usage_temp': order.usage_temp,
        'quantity': order.quantity,
        'deadline': order.deadline,
        'location': order.location,
        'description': order.description,
        'file_path': order.file_path,
        'language': order.language,
        'manager_contact': order.manager_contact,
        'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
        'updated_at': order.updated_at.strftime('%d.%m.%Y %H:%M')
    }
    
    return render_template('order_detail.html', 
                         order=order_data,
                         statuses=Config.ORDER_STATUSES)

@app.route('/admin/order/<int:order_id>/update', methods=['POST'])
def update_order(order_id):
    """Update order status and details"""
    try:
        order = db.get_order(order_id)
        if not order:
            flash('Заказ не найден', 'error')
            return redirect(url_for('admin'))
        
        # Update order data
        new_status = request.form.get('status')
        manager_contact = request.form.get('manager_contact', '')
        
        # Update in database
        updated_order = db.update_order_status(order_id, new_status, manager_contact)
        
        if updated_order:
            flash('Заказ успешно обновлен', 'success')
        else:
            flash('Ошибка при обновлении заказа', 'error')
        
        return redirect(url_for('order_detail', order_id=order_id))
    
    except Exception as e:
        flash(f'Ошибка при обновлении заказа: {str(e)}', 'error')
        return redirect(url_for('order_detail', order_id=order_id))

@app.route('/admin/export')
def export_orders():
    """Export orders to Excel"""
    try:
        # Get filter parameters
        status_filter = request.args.get('status', 'all')
        
        # Get orders
        if status_filter == 'all':
            orders = db.get_all_orders()
        else:
            orders = db.get_orders_by_status(status_filter)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        # Headers
        headers = [
            'ID',
            'Имя клиента',
            'Телефон',
            'Username',
            'Тип услуги',
            'Статус',
            'Материал',
            'Цвет',
            'Количество',
            'Срок',
            'Локация',
            'Описание',
            'Дата создания',
            'Менеджер'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.id)
            ws.cell(row=row, column=2, value=order.name)
            ws.cell(row=row, column=3, value=order.phone)
            ws.cell(row=row, column=4, value=order.username)
            ws.cell(row=row, column=5, value=Config.SERVICE_TYPES.get(order.service_type, order.service_type))
            ws.cell(row=row, column=6, value=Config.ORDER_STATUSES.get(order.status, order.status))
            ws.cell(row=row, column=7, value=order.material)
            ws.cell(row=row, column=8, value=order.color)
            ws.cell(row=row, column=9, value=order.quantity)
            ws.cell(row=row, column=10, value=order.deadline)
            ws.cell(row=row, column=11, value=order.location)
            ws.cell(row=row, column=12, value=order.description)
            ws.cell(row=row, column=13, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=14, value=order.manager_contact)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            BytesIO(output.getvalue()),
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        flash(f'Ошибка при экспорте: {str(e)}', 'error')
        return redirect(url_for('admin'))

@app.route('/admin/api/orders')
def api_orders():
    """API endpoint for orders"""
    try:
        status_filter = request.args.get('status', 'all')
        
        if status_filter == 'all':
            orders = db.get_all_orders()
        else:
            orders = db.get_orders_by_status(status_filter)
        
        orders_data = []
        for order in orders:
            orders_data.append({
                'id': order.id,
                'name': order.name,
                'phone': order.phone,
                'service_type': order.service_type,
                'status': order.status,
                'created_at': order.created_at.isoformat()
            })
        
        return jsonify(orders_data)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/api/order/<int:order_id>/status', methods=['PUT'])
def api_update_order_status(order_id):
    """API endpoint to update order status"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'error': 'Status is required'}), 400
        
        updated_order = db.update_order_status(order_id, new_status)
        
        if updated_order:
            return jsonify({
                'success': True,
                'order_id': order_id,
                'status': new_status
            })
        else:
            return jsonify({'error': 'Order not found'}), 404
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(
        host=Config.FLASK_HOST,
        port=Config.FLASK_PORT,
        debug=Config.DEBUG
    )
